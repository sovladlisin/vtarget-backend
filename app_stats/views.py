from json.decoder import JSONDecodeError
from django.http.response import JsonResponse
from django.shortcuts import render
from think_bank.views import vk_request
from django.http import StreamingHttpResponse, HttpResponseRedirect, HttpResponse
import datetime
from .models import AppsIds
from users.models import VkUser
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
# Create your views here.
import json


def analyzeExcel(data_file):
    result = []
    wb = load_workbook(data_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws:
            if not any(cell.value for cell in row):
                pass
            else:
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        try:
                            val = int(cell.value)
                            if len(str(val)) == 7:
                                result.append(val)
                        except:
                            pass
                    elif cell.value and (isinstance(cell.value, float) or isinstance(cell.value, int)) and (len(str(int(cell.value)))) == 7:
                        result.append(int(cell.value))
    return result


@csrf_exempt
def getExcelIds(request):
    if request.method == 'POST':
        file_d = request.FILES['file']
        excel_file = file_d.read()
        excel_ids = analyzeExcel(excel_file)
        return JsonResponse(excel_ids, safe=False)


@csrf_exempt
def getAppList(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        number_of_elements = data.get('number_of_elements', 0)
        page_number = data.get('page_number', 0)
        search = data.get('search', '')

        apps_ids = AppsIds.objects.all().first()
        data_apps = json.loads(apps_ids.data)

        # data_apps_searched = []
        # for i in data_apps:
        #     if search in i['title']:
        #         data_apps_searched.append(i)

        sorted_data = sorted(
            data_apps, key=lambda x: x['members_count'], reverse=True)
        # chunks_data = chunks(sorted_data, number_of_elements)

        return JsonResponse({'data': sorted_data, 'size': len(sorted_data)})


@csrf_exempt
def updateIds(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        ids = data.get('ids', None)
        if ids is None:
            return HttpResponse(status=403)
        apps_ids = AppsIds.objects.all().first()
        apps_ids.ids = json.dumps(ids)
        apps_ids.save()
        return HttpResponse(status=200)


def getAppAnalyzeStatus(request):
    if request.method == 'GET':
        result = {}
        apps_ids = AppsIds.objects.all().first()
        result['ids'] = json.loads(apps_ids.ids)
        result['progress'] = apps_ids.progress
        result['in_progress'] = apps_ids.in_progress
        result['last_updated'] = apps_ids.last_updated

        return JsonResponse(result, safe=False)


class ResponseThen(HttpResponse):
    def __init__(self, data, then_callback, token, **kwargs):
        super().__init__(data, **kwargs)
        self.then_callback = then_callback
        self.token = token

    def close(self):
        super().close()
        self.then_callback(self.token)


@csrf_exempt
def analyzeApps(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        user_id = data.get('user_id', None)
        user = VkUser.objects.get(pk=user_id)

        apps_ids = AppsIds.objects.all().first()
        apps_ids.last_updated = datetime.datetime.now()

        return ResponseThen('ok', processIds, user.token, status=200)


def processIds(token):
    result = []
    ids_object = AppsIds.objects.all().first()
    ids_object.in_progress = True
    ids_object.save()
    ids = json.loads(ids_object.ids)

    step = 0
    for id in ids:
        step += 1
        ids_object.progress = str(step / len(ids))
        ids_object.save()
        print(id)
        info = get_app_info(id, token)
        if info.get('id', None) is not None:
            result.append(get_app_info(id, token))

    ids_object.in_progress = False
    ids_object.progress = '100'
    ids_object.data = json.dumps(result)
    ids_object.save()
    return True


def get_app_info(id, token):
    return vk_request('get', 'apps.get', {'app_id': id}, token, '5.21')['response']


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]
