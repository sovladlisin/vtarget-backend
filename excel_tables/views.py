from json.decoder import JSONDecodeError
from django.shortcuts import render
from django.http import StreamingHttpResponse, HttpResponseRedirect, HttpResponse
from django.forms.models import model_to_dict
from django.http import JsonResponse
from openpyxl import load_workbook
from users.models import VkUser
from .models import ExcelTable
from django.views.decorators.csrf import csrf_exempt
import json
# Create your views here.


@csrf_exempt
def uploadExcelTable(request):
    if request.method == 'POST':
        user_pk = request.GET['user_pk']
        is_public = request.GET['is_public']
        title = request.GET['title']
        user = VkUser.objects.get(pk=user_pk)

        file_d = request.FILES['file']
        excel_file = file_d
        tables = buildTablesFromExcel(excel_file)

        response = []
        is_public_local = True if int(is_public) == 1 else False
        for table in tables:
            new_table = ExcelTable(
                owner=user, data=json.dumps(table), is_public=is_public_local, title=title)
            new_table.save()
            response.append(
                {'id': new_table.pk, 'is_public': new_table.is_public, 'owner': model_to_dict(new_table.owner)})
        return JsonResponse(response, safe=False)


@csrf_exempt
def updateExcelTable(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        title = data.get('title', '')
        id = data.get('id', -1)
        is_public = data.get('is_public', False)

        table = ExcelTable.objects.get(pk=id)
        table.is_public = is_public
        table.title = title
        table.save()
        return HttpResponse(status=200)


@csrf_exempt
def getExcelTables(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        user_pk = data.get('user_pk', -1)

        public = ExcelTable.objects.all().filter(
            is_public=True).exclude(owner__pk=user_pk)
        owned = ExcelTable.objects.all().filter(owner__pk=user_pk)

        response = {'public': [], 'owned': []}

        for t in public:
            response['public'].append(
                {'id': t.pk, 'is_public': t.is_public, 'owner': model_to_dict(t.owner)})

        for t in owned:
            response['owned'].append(
                {'id': t.pk, 'is_public': t.is_public, 'owner': model_to_dict(t.owner)})

        return JsonResponse(response, safe=False)


@csrf_exempt
def getExcelTable(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        user_pk = data.get('user_pk', -1)
        table_pk = data.get('table_pk', -1)

        t = ExcelTable.objects.get(pk=table_pk)
        if t.is_public == False and t.owner.pk != user_pk:
            return HttpResponse(status=403)

        response = {'id': t.pk, 'is_public': t.is_public,
                    'owner': model_to_dict(t.owner), 'data': json.loads(t.data)}

        return JsonResponse(response, safe=False)


@csrf_exempt
def deleteExcelTable(request):
    if request.method == 'DELETE':
        id = request.GET['id']
        table = ExcelTable.objects.get(pk=id)
        table.delete()
        return HttpResponse(status=200)


def buildTablesFromExcel(data_file):
    result = []
    wb = load_workbook(data_file)

    for sheet in wb.sheetnames:
        count = 0
        ws = wb[sheet]
        table = {
            'table_names': ['ID'],
            'rows': []
        }
        for row in ws:
            if not any(cell.value for cell in row):
                pass
            else:
                count += 1
                if count == 1:
                    for cell in row:
                        table['table_names'].append(str(cell.value))

                else:
                    result_row = {}
                    cell_count_max = len(table['table_names'])
                    cell_count = 0
                    for cell in row:
                        cell_count += 1
                        if cell_count > cell_count_max:
                            pass
                        else:
                            value = '' if cell.value is None else cell.value
                            index = table['table_names'][cell_count-1]
                            result_row[index] = value
                    result_row['id'] = count - 2
                    table['rows'].append(result_row)
        result.append(table)
    return result
